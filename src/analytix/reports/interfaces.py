# SPDX-FileCopyrightText: 2021-2026 Ethan Henderson
# SPDX-FileCopyrightText: 2022 Jonxslays
# SPDX-License-Identifier: BSD-3-Clause

"""Report interfaces for analytix.

These are report interfaces equipped with various methods of saving and
exporting report data to different formats. They are not designed to be
like-for-like mappings of YouTube Analytics API resources.

Currently, there is only one of these interfaces.
"""

__all__ = ("Report",)

import json
import logging
from io import TextIOBase
from os import PathLike
from pathlib import Path
from typing import TYPE_CHECKING
from typing import Any

from analytix import utils
from analytix.errors import DataFrameConversionError
from analytix.errors import MissingOptionalComponents
from analytix.reports.resources import ColumnType
from analytix.reports.resources import ResultTable

if TYPE_CHECKING:
    import pandas as pd
    import polars as pl

    from analytix.abc import ReportType

_log = logging.getLogger(__name__)


class Report:
    """An analytics report.

    This is an abstraction of the `resultTable` resource rather than a
    direct mapping. This class provides additional properties and
    methods designed to make it easier to perform certain operations.

    ???+ note "Changed in version 5.0"
        This used to be `AnalyticsReport`.

    Parameters
    ----------
    data
        The raw JSON data from the API.
    type
        The report type.

    Attributes
    ----------
    resource : ResultTable
        An instance representing a `resultTable` resource.
    type : ReportType
        The report type.
    """

    def __init__(self, data: dict[str, Any], type: "ReportType") -> None:
        self.resource = ResultTable.from_json(data)
        self.type = type

    @property
    def shape(self) -> tuple[int, int]:
        """The shape of the report.

        This is presented in (rows, columns) format.

        Returns
        -------
        Tuple[int, int]
            The shape of the report.

        Examples
        --------
        >>> report.shape
        (120, 42)
        """
        return (len(self.resource.rows), len(self.resource.column_headers))

    @property
    def columns(self) -> list[str]:
        """A list of all columns names in the report.

        Returns
        -------
        List[str]
            The column list.

        See Also
        --------
        This does not return a list of column headers. If you want that,
        use `report.resource.column_headers` instead.

        Examples
        --------
        >>> report.columns
        ["day", "subscribedStatus", "views", "likes", "comments"]
        """
        return [c.name for c in self.resource.column_headers]

    @property
    def dimensions(self) -> list[str]:
        """A list of all dimensions in the report.

        Returns
        -------
        List[str]
            The dimension list.

        Examples
        --------
        >>> report.dimensions
        ["day", "subscribedStatus"]
        """
        return [
            c.name
            for c in self.resource.column_headers
            if c.column_type == ColumnType.DIMENSION
        ]

    @property
    def metrics(self) -> list[str]:
        """A list of all metrics in the report.

        Returns
        -------
        List[str]
            The metric list.

        Examples
        --------
        >>> report.metrics
        ["views", "likes", "comments"]
        """
        return [
            c.name
            for c in self.resource.column_headers
            if c.column_type == ColumnType.METRIC
        ]

    def to_json(
        self,
        path_or_buf: str | PathLike[str] | TextIOBase | None = None,
        *,
        overwrite: bool = False,
        **kwargs: Any,
    ) -> str | None:
        """Convert this report to a JSON string.

        This saves the data as it arrived from the YouTube Analytics
        API.

        Parameters
        ----------
        path_of_buf
            The path to save the report to, a file-like object to write
            to, or `None`.
        overwrite
            Whether to overwrite an existing file.
        **kwargs
            Additional arguments to pass to `json.dumps()`

        Returns
        -------
        str | None
            The JSON string if `path_or_buf` is `None`, otherwise
            `None`.

        Examples
        --------
        >>> report.to_json()
        {"kind": "youtubeAnalytics#resultTable", ...}

        Writing to a file.

        >>> report.to_json("report.json", indent=4)

        Writing to a file-like object.

        >>> report.to_json(buf := io.StringIO())
        >>> buf.getvalue()
        '{"kind": "youtubeAnalytics#resultTable", ...}'
        """
        if path_or_buf is None:
            return json.dumps(self.resource.data, **kwargs)

        if isinstance(path_or_buf, (str, PathLike)):
            output_path = Path(path_or_buf)
            if (not overwrite) and output_path.exists():
                raise FileExistsError(
                    "file already exists and `overwrite` is set to False",
                )
            output_path.write_text(json.dumps(self.resource.data, **kwargs))
            _log.debug("Saved report to %s", output_path.resolve())
            return None

        if isinstance(path_or_buf, TextIOBase):
            path_or_buf.write(json.dumps(self.resource.data, **kwargs))
            return None

        raise TypeError(
            f"Expected str, PathLike, or TextIOBase, got {type(path_or_buf).__name__}",
        )

    def to_csv(
        self,
        path_or_buf: str | PathLike[str] | TextIOBase | None = None,
        *,
        delimiter: str = ",",
        overwrite: bool = False,
    ) -> str | None:
        """Convert this report to a CSV or TSV string.

        Parameters
        ----------
        path_of_buf
            The path to save the report to, a file-like object to write
            to, or `None`.
        delimiter
            The character to use as a delimiter.
        overwrite
            Whether to overwrite an existing file.

        Returns
        -------
        str | None
            The CSV string if `path_or_buf` is `None`, otherwise `None`.

        Examples
        --------
        >>> report.to_csv()
        'views,redViews,comments,likes,dislikes,...'

        Writing to a file.

        >>> report.to_csv("report.csv")

        Writing to a file-like object.

        >>> report.to_csv(buf := io.StringIO())
        >>> buf.getvalue()
        'views,redViews,comments,likes,dislikes,...'
        """
        output = f"{delimiter.join(self.columns)}\n"
        for row in self.resource.rows:
            line = delimiter.join(f"{v}" for v in row)
            output += f"{line}\n"

        if path_or_buf is None:
            return output

        if isinstance(path_or_buf, (str, PathLike)):
            output_path = Path(path_or_buf)
            if (not overwrite) and output_path.exists():
                raise FileExistsError(
                    "file already exists and `overwrite` is set to False",
                )
            output_path.write_text(output)
            _log.debug("Saved report to %s", output_path.resolve())
            return None

        if isinstance(path_or_buf, TextIOBase):
            path_or_buf.write(output)
            return None

        raise TypeError(
            f"Expected str, PathLike, or TextIOBase, got {type(path_or_buf).__name__}",
        )

    def to_excel(
        self,
        path: str | PathLike[str],
        *,
        sheet_name: str = "Analytics",
    ) -> None:
        """Save this report as an Excel spreadsheet.

        If a workbook already exists in the given path, the data will be
        inserted into a new sheet with the given name.

        Parameters
        ----------
        path
            The path to save the spreadsheet to.
        sheet_name
            The name to give the sheet the data will be inserted into.

        Notes
        -----
        This requires `openpyxl` to be installed to use, which is an
        optional dependency.

        Examples
        --------
        >>> report.to_excel("output.xlsx")
        """
        if not utils.can_use("openpyxl"):
            raise MissingOptionalComponents("openpyxl")

        if not isinstance(path, (str, PathLike)):
            raise TypeError(f"Expected str or PathLike, got {type(path).__name__}")

        output_path = Path(path)
        if output_path.exists():
            _log.debug("Workbook already exists; adding new sheet")
            from openpyxl import load_workbook

            wb = load_workbook(output_path)
            ws = wb.create_sheet(sheet_name)
        else:
            _log.debug("Workbook does not exist; creating new workbook")
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active or wb.create_sheet()
            ws.title = sheet_name

        ws.append(self.columns)
        for row in self.resource.rows:
            ws.append(row)

        wb.save(str(output_path))
        _log.debug("Saved report to %s", output_path.resolve())

    def to_pandas(self) -> "pd.DataFrame":
        """Return this report as a pandas DataFrame.

        Parameters
        ----------
        skip_date_conversion
            Whether or not to skip the conversion of "day" and "month"
            columns into a datetime format. If you choose to skip this,
            these columns will be left as strings.

        Returns
        -------
        pandas DataFrame
            A pandas DataFrame.

        Raises
        ------
        MissingOptionalComponents
            pandas is not installed.
        DataFrameConversionError
            There is no data from which to create a DataFrame.

        Notes
        -----
        This requires `pandas` to be installed to use, which is an
        optional dependency.

        Examples
        --------
        >>> df = report.to_pandas()
        >>> df.head(5)
                 day  views  likes  comments  grossRevenue
        0 2022-06-20    778      8         0         2.249
        1 2022-06-21   1062     32         8         3.558
        2 2022-06-22    946     38         6         2.910
        3 2022-06-23   5107    199        15        24.428
        4 2022-06-24   2137     61         2         6.691
        """
        # sourcery skip: class-extract-method
        if not utils.can_use("pandas"):
            raise MissingOptionalComponents("pandas")

        if not self.shape[0]:
            raise DataFrameConversionError(
                "cannot convert to DataFrame as the returned data has no rows",
            )

        import pandas as pd

        df = pd.DataFrame(self.resource.rows, columns=self.columns)

        if len(s := {"day", "month"} & set(df.columns)):
            col = next(iter(s))
            fmt = {"day": "%Y-%m-%d", "month": "%Y-%m"}[col]
            df[col] = pd.to_datetime(df[col], format=fmt)
            _log.debug(f"Converted {col!r} column to datetime format")

        return df

    def to_polars(self) -> "pl.DataFrame":
        """Return the data as a Polars DataFrame.

        Parameters
        ----------
        skip_date_conversion
            Whether or not to skip the conversion of "day" and "month"
            columns into a date format. If you choose to skip this,
            these columns will be left as strings.

        Returns
        -------
        Polars DataFrame
            A Polars DataFrame.

        Raises
        ------
        MissingOptionalComponents
            Polars is not installed.
        DataFrameConversionError
            There is no data from which to create a DataFrame.

        Notes
        -----
        This requires `polars` to be installed to use, which is an
        optional dependency.

        Examples
        --------
        >>> df = report.to_polars()
        >>> df.head(5)
        shape: (5, 5)
        ┌────────────┬───────┬───────┬──────────┬──────────────┐
        │ day        ┆ views ┆ likes ┆ comments ┆ grossRevenue │
        │ ---        ┆ ---   ┆ ---   ┆ ---      ┆ ---          │
        │ date       ┆ i64   ┆ i64   ┆ i64      ┆ f64          │
        ╞════════════╪═══════╪═══════╪══════════╪══════════════╡
        │ 2022-06-20 ┆ 778   ┆ 8     ┆ 0        ┆ 2.249        │
        ├╌╌╌╌╌╌╌╌╌╌╌╌┼╌╌╌╌╌╌╌┼╌╌╌╌╌╌╌┼╌╌╌╌╌╌╌╌╌╌┼╌╌╌╌╌╌╌╌╌╌╌╌╌╌┤
        │ 2022-06-21 ┆ 1062  ┆ 32    ┆ 8        ┆ 3.558        │
        ├╌╌╌╌╌╌╌╌╌╌╌╌┼╌╌╌╌╌╌╌┼╌╌╌╌╌╌╌┼╌╌╌╌╌╌╌╌╌╌┼╌╌╌╌╌╌╌╌╌╌╌╌╌╌┤
        │ 2022-06-22 ┆ 946   ┆ 38    ┆ 6        ┆ 2.91         │
        ├╌╌╌╌╌╌╌╌╌╌╌╌┼╌╌╌╌╌╌╌┼╌╌╌╌╌╌╌┼╌╌╌╌╌╌╌╌╌╌┼╌╌╌╌╌╌╌╌╌╌╌╌╌╌┤
        │ 2022-06-23 ┆ 5107  ┆ 199   ┆ 15       ┆ 24.428       │
        ├╌╌╌╌╌╌╌╌╌╌╌╌┼╌╌╌╌╌╌╌┼╌╌╌╌╌╌╌┼╌╌╌╌╌╌╌╌╌╌┼╌╌╌╌╌╌╌╌╌╌╌╌╌╌┤
        │ 2022-06-24 ┆ 2137  ┆ 61    ┆ 2        ┆ 6.691        │
        └────────────┴───────┴───────┴──────────┴──────────────┘
        """
        if not utils.can_use("polars"):
            raise MissingOptionalComponents("polars")

        if not self.shape[0]:
            raise DataFrameConversionError(
                "cannot convert to DataFrame as the returned data has no rows",
            )

        import polars as pl

        df = pl.DataFrame(self.resource.rows, schema=self.columns)

        if len(s := {"day", "month"} & set(df.columns)):
            col = next(iter(s))
            fmt = {"day": "%Y-%m-%d", "month": "%Y-%m"}[col]
            df = df.with_columns(pl.col(col).str.strptime(pl.Date, fmt))
            _log.debug(f"Converted {col!r} column to date format")

        return df
