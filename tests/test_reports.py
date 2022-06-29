# Copyright (c) 2021-present, Ethan Henderson
# All rights reserved.
#
# Redistribution and use in source and binary forms, with or without
# modification, are permitted provided that the following conditions are met:
#
# 1. Redistributions of source code must retain the above copyright notice, this
#    list of conditions and the following disclaimer.
#
# 2. Redistributions in binary form must reproduce the above copyright notice,
#    this list of conditions and the following disclaimer in the documentation
#    and/or other materials provided with the distribution.
#
# 3. Neither the name of the copyright holder nor the names of its
#    contributors may be used to endorse or promote products derived from
#    this software without specific prior written permission.
#
# THIS SOFTWARE IS PROVIDED BY THE COPYRIGHT HOLDERS AND CONTRIBUTORS "AS IS"
# AND ANY EXPRESS OR IMPLIED WARRANTIES, INCLUDING, BUT NOT LIMITED TO, THE
# IMPLIED WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE ARE
# DISCLAIMED. IN NO EVENT SHALL THE COPYRIGHT HOLDER OR CONTRIBUTORS BE LIABLE
# FOR ANY DIRECT, INDIRECT, INCIDENTAL, SPECIAL, EXEMPLARY, OR CONSEQUENTIAL
# DAMAGES (INCLUDING, BUT NOT LIMITED TO, PROCUREMENT OF SUBSTITUTE GOODS OR
# SERVICES; LOSS OF USE, DATA, OR PROFITS; OR BUSINESS INTERRUPTION) HOWEVER
# CAUSED AND ON ANY THEORY OF LIABILITY, WHETHER IN CONTRACT, STRICT LIABILITY,
# OR TORT (INCLUDING NEGLIGENCE OR OTHERWISE) ARISING IN ANY WAY OUT OF THE USE
# OF THIS SOFTWARE, EVEN IF ADVISED OF THE POSSIBILITY OF SUCH DAMAGE.

import json
import os
import platform
import sys

import mock
import pytest

import analytix
from analytix import data, errors
from analytix.report_types import TimeBasedActivity
from analytix.reports import (
    ColumnHeader,
    ColumnType,
    DataType,
    Report,
)
from tests.paths import (
    CSV_OUTPUT_PATH,
    EXCEL_OUTPUT_PATH,
    FEATHER_OUTPUT_PATH,
    JSON_OUTPUT_PATH,
    MOCK_CSV_PATH,
    MOCK_DATA_PATH,
    PARQUET_OUTPUT_PATH,
    TSV_OUTPUT_PATH,
)

if analytix.can_use("openpyxl"):
    from openpyxl import load_workbook

if analytix.can_use("pandas"):
    import pandas as pd


@pytest.fixture()
def request_data():
    with open(MOCK_DATA_PATH) as f:
        return json.load(f)


@pytest.fixture()
def report_type():
    return TimeBasedActivity()


def test_init(request_data, report_type):
    report = Report(request_data, report_type)
    assert report.data == request_data
    assert report.type == report_type
    assert report._column_headers == [
        ColumnHeader("day", ColumnType.DIMENSION, DataType.STRING),
        ColumnHeader("views", ColumnType.METRIC, DataType.INTEGER),
        ColumnHeader("redViews", ColumnType.METRIC, DataType.INTEGER),
        ColumnHeader("comments", ColumnType.METRIC, DataType.INTEGER),
        ColumnHeader("likes", ColumnType.METRIC, DataType.INTEGER),
        ColumnHeader("dislikes", ColumnType.METRIC, DataType.INTEGER),
        ColumnHeader("videosAddedToPlaylists", ColumnType.METRIC, DataType.INTEGER),
        ColumnHeader("videosRemovedFromPlaylists", ColumnType.METRIC, DataType.INTEGER),
        ColumnHeader("shares", ColumnType.METRIC, DataType.INTEGER),
        ColumnHeader("estimatedMinutesWatched", ColumnType.METRIC, DataType.INTEGER),
        ColumnHeader("estimatedRedMinutesWatched", ColumnType.METRIC, DataType.INTEGER),
        ColumnHeader("averageViewDuration", ColumnType.METRIC, DataType.INTEGER),
        ColumnHeader("averageViewPercentage", ColumnType.METRIC, DataType.FLOAT),
        ColumnHeader("annotationClickThroughRate", ColumnType.METRIC, DataType.FLOAT),
        ColumnHeader("annotationCloseRate", ColumnType.METRIC, DataType.FLOAT),
        ColumnHeader("annotationImpressions", ColumnType.METRIC, DataType.INTEGER),
        ColumnHeader(
            "annotationClickableImpressions", ColumnType.METRIC, DataType.INTEGER
        ),
        ColumnHeader(
            "annotationClosableImpressions", ColumnType.METRIC, DataType.INTEGER
        ),
        ColumnHeader("annotationClicks", ColumnType.METRIC, DataType.INTEGER),
        ColumnHeader("annotationCloses", ColumnType.METRIC, DataType.INTEGER),
        ColumnHeader("cardClickRate", ColumnType.METRIC, DataType.FLOAT),
        ColumnHeader("cardTeaserClickRate", ColumnType.METRIC, DataType.FLOAT),
        ColumnHeader("cardImpressions", ColumnType.METRIC, DataType.INTEGER),
        ColumnHeader("cardTeaserImpressions", ColumnType.METRIC, DataType.INTEGER),
        ColumnHeader("cardClicks", ColumnType.METRIC, DataType.INTEGER),
        ColumnHeader("cardTeaserClicks", ColumnType.METRIC, DataType.INTEGER),
        ColumnHeader("subscribersGained", ColumnType.METRIC, DataType.INTEGER),
        ColumnHeader("subscribersLost", ColumnType.METRIC, DataType.INTEGER),
        ColumnHeader("estimatedRevenue", ColumnType.METRIC, DataType.FLOAT),
        ColumnHeader("estimatedAdRevenue", ColumnType.METRIC, DataType.FLOAT),
        ColumnHeader("grossRevenue", ColumnType.METRIC, DataType.FLOAT),
        ColumnHeader("estimatedRedPartnerRevenue", ColumnType.METRIC, DataType.FLOAT),
        ColumnHeader("monetizedPlaybacks", ColumnType.METRIC, DataType.INTEGER),
        ColumnHeader("playbackBasedCpm", ColumnType.METRIC, DataType.FLOAT),
        ColumnHeader("adImpressions", ColumnType.METRIC, DataType.INTEGER),
        ColumnHeader("cpm", ColumnType.METRIC, DataType.FLOAT),
    ]
    assert report._shape == (31, 36)


@pytest.fixture()
def report():
    with open(MOCK_DATA_PATH) as f:
        data = json.load(f)

    return Report(data, TimeBasedActivity())


def test_shape_property(report):
    assert report.shape == (31, 36)


def test_rows_property(report):
    # There's a HUGE number here, so testing the first row should be
    # sufficient to say it's correct.
    assert report.rows[0] == [
        "2022-01-01",
        759,
        82,
        7,
        15,
        0,
        7,
        7,
        0,
        1335,
        68,
        105,
        7.01,
        0,
        0,
        0,
        0,
        0,
        0,
        0,
        0,
        0,
        0,
        14,
        0,
        0,
        3,
        4,
        1.096,
        1.081,
        1.965,
        0.015,
        198,
        9.924,
        239,
        8.222,
    ]


def test_column_headers_property(report):
    assert report.column_headers == report._column_headers


def test_columns_property(report):
    assert report.columns == [
        "day",
        *[m for m in data.ALL_METRICS_ORDERED if m in report.type.metrics.values],
    ]


def test_dimensions_property(report):
    assert report.dimensions == {"day"}


def test_metrics_property(report):
    assert report.metrics == data.ALL_VIDEO_METRICS


def test_ordered_dimensions_property(report):
    assert report.ordered_dimensions == ["day"]


def test_ordered_metrics_property(report):
    assert report.ordered_metrics == [
        m for m in data.ALL_METRICS_ORDERED if m in report.type.metrics.values
    ]


def test_to_json(report, request_data):
    report.to_json(str(JSON_OUTPUT_PATH))
    assert JSON_OUTPUT_PATH.is_file()

    with open(JSON_OUTPUT_PATH) as f:
        assert json.load(f) == request_data

    os.remove(JSON_OUTPUT_PATH)


def test_to_json_no_extension(report, request_data):
    report.to_json(str(JSON_OUTPUT_PATH)[:-5])
    assert JSON_OUTPUT_PATH.is_file()

    with open(JSON_OUTPUT_PATH) as f:
        assert json.load(f) == request_data

    os.remove(JSON_OUTPUT_PATH)


@pytest.fixture()
def mock_csv_data():
    with open(MOCK_CSV_PATH) as f:
        return f.read()


def test_to_csv(report, mock_csv_data):
    report.to_csv(str(CSV_OUTPUT_PATH))
    assert CSV_OUTPUT_PATH.is_file()

    with open(CSV_OUTPUT_PATH) as f:
        assert f.read() == mock_csv_data

    os.remove(CSV_OUTPUT_PATH)


def test_to_csv_no_extension(report, mock_csv_data):
    report.to_csv(str(CSV_OUTPUT_PATH)[:-4])
    assert CSV_OUTPUT_PATH.is_file()

    with open(CSV_OUTPUT_PATH) as f:
        assert f.read() == mock_csv_data

    os.remove(CSV_OUTPUT_PATH)


def test_to_tsv(report, mock_csv_data):
    report.to_csv(str(TSV_OUTPUT_PATH), delimiter="\t")
    assert TSV_OUTPUT_PATH.is_file()

    with open(TSV_OUTPUT_PATH) as f:
        assert f.read() == mock_csv_data.replace(",", "\t")

    os.remove(TSV_OUTPUT_PATH)


def test_to_tsv_no_extension(report, mock_csv_data):
    report.to_csv(str(TSV_OUTPUT_PATH)[:-4], delimiter="\t")
    assert TSV_OUTPUT_PATH.is_file()

    with open(TSV_OUTPUT_PATH) as f:
        assert f.read() == mock_csv_data.replace(",", "\t")

    os.remove(TSV_OUTPUT_PATH)


@pytest.mark.skipif(
    sys.version_info >= (3, 11, 0) or platform.python_implementation() != "CPython",
    reason="pandas does not support Python 3.11 or PyPy",
)
def test_to_dataframe(report, request_data):
    df = report.to_dataframe()
    assert df.shape == (31, 36)
    assert list(df.columns) == report.columns

    for i, row in df.iterrows():
        # We don't need the date to check -- it just complicated things.
        assert list(row)[1:] == request_data["rows"][i][1:]


@pytest.mark.skipif(
    sys.version_info >= (3, 11, 0) or platform.python_implementation() != "CPython",
    reason="pandas does not support Python 3.11 or PyPy",
)
def test_to_dataframe_modin_check(report):
    # It's not really possible to test Modin functionality, so just run
    # this to check Modin can be selected, and get the coverage.

    with mock.patch.object(analytix, "can_use") as mock_cu:
        mock_cu.return_value = True

        with pytest.raises(ImportError):
            df = report.to_dataframe()


@pytest.mark.skipif(
    sys.version_info >= (3, 11, 0) or platform.python_implementation() != "CPython",
    reason="pandas does not support Python 3.11 or PyPy",
)
def test_to_dataframe_no_pandas(report):
    with mock.patch.object(analytix, "can_use") as mock_cu:
        mock_cu.return_value = False

        with pytest.raises(errors.MissingOptionalComponents) as exc:
            report.to_dataframe()
        assert (
            str(exc.value)
            == "some necessary libraries are not installed (hint: pip install pandas)"
        )


@pytest.mark.skipif(
    sys.version_info >= (3, 11, 0) or platform.python_implementation() != "CPython",
    reason="pandas does not support Python 3.11 or PyPy",
)
def test_to_dataframe_no_rows(report):
    report._shape = (0, 0)

    with pytest.raises(errors.DataFrameConversionError) as exc:
        report.to_dataframe()
    assert (
        str(exc.value) == "cannot convert to DataFrame as the returned data has no rows"
    )


def test_to_excel(report, mock_csv_data):
    report.to_excel(str(EXCEL_OUTPUT_PATH))
    assert EXCEL_OUTPUT_PATH.is_file()

    ws = load_workbook(EXCEL_OUTPUT_PATH)["Analytics"]
    excel_data = "\n".join(",".join(f"{cell.value}" for cell in row) for row in ws.rows)
    assert excel_data == mock_csv_data.strip()

    try:
        os.remove(EXCEL_OUTPUT_PATH)
    except PermissionError:
        # Account for bizarre PermissionError on Windows PyPy tests.
        ...


def test_to_excel_no_extension(report, mock_csv_data):
    report.to_excel(str(EXCEL_OUTPUT_PATH)[:-5])
    assert EXCEL_OUTPUT_PATH.is_file()

    ws = load_workbook(EXCEL_OUTPUT_PATH)["Analytics"]
    excel_data = "\n".join(",".join(f"{cell.value}" for cell in row) for row in ws.rows)
    assert excel_data == mock_csv_data.strip()

    try:
        os.remove(EXCEL_OUTPUT_PATH)
    except PermissionError:
        # Account for bizarre PermissionError on Windows PyPy tests.
        ...


def test_to_excel_no_openpyxl(report):
    with mock.patch.object(analytix, "can_use") as mock_cu:
        mock_cu.return_value = False

        with pytest.raises(errors.MissingOptionalComponents) as exc:
            report.to_excel(EXCEL_OUTPUT_PATH)
        assert (
            str(exc.value)
            == "some necessary libraries are not installed (hint: pip install openpyxl)"
        )


@pytest.mark.skipif(
    sys.version_info >= (3, 11, 0) or platform.python_implementation() != "CPython",
    reason="PyArrow does not support Python 3.11",
)
def test_to_arrow_table(report):
    table = report.to_arrow_table()

    df_arrow = table.to_pandas()
    df_csv = pd.read_csv(MOCK_CSV_PATH)
    df_csv["day"] = pd.to_datetime(df_csv["day"], format="%Y-%m-%d")
    assert df_arrow.equals(df_csv)


@pytest.mark.skipif(
    sys.version_info >= (3, 11, 0) or platform.python_implementation() != "CPython",
    reason="PyArrow does not support Python 3.11",
)
def test_to_arrow_table_no_pyarrow(report):
    with mock.patch.object(analytix, "can_use") as mock_cu:
        mock_cu.return_value = False

        with pytest.raises(errors.MissingOptionalComponents) as exc:
            report.to_arrow_table()
        assert (
            str(exc.value)
            == "some necessary libraries are not installed (hint: pip install pyarrow)"
        )


@pytest.mark.skipif(
    sys.version_info >= (3, 11, 0) or platform.python_implementation() != "CPython",
    reason="PyArrow does not support Python 3.11",
)
def test_to_feather(report):
    report.to_feather(str(FEATHER_OUTPUT_PATH))
    assert FEATHER_OUTPUT_PATH.is_file()

    df_csv = pd.read_csv(MOCK_CSV_PATH)
    df_csv["day"] = pd.to_datetime(df_csv["day"], format="%Y-%m-%d")
    df_feather = pd.read_feather(FEATHER_OUTPUT_PATH)
    assert df_feather.equals(df_csv)

    os.remove(FEATHER_OUTPUT_PATH)


@pytest.mark.skipif(
    sys.version_info >= (3, 11, 0) or platform.python_implementation() != "CPython",
    reason="PyArrow does not support Python 3.11",
)
def test_to_feather_no_extension(report):
    report.to_feather(str(FEATHER_OUTPUT_PATH)[:-8])
    assert FEATHER_OUTPUT_PATH.is_file()

    df_csv = pd.read_csv(MOCK_CSV_PATH)
    df_csv["day"] = pd.to_datetime(df_csv["day"], format="%Y-%m-%d")
    df_feather = pd.read_feather(FEATHER_OUTPUT_PATH)
    assert df_feather.equals(df_csv)

    os.remove(FEATHER_OUTPUT_PATH)


@pytest.mark.skipif(
    sys.version_info >= (3, 11, 0) or platform.python_implementation() != "CPython",
    reason="PyArrow does not support Python 3.11",
)
def test_to_feather_no_pyarrow(report):
    with mock.patch.object(analytix, "can_use") as mock_cu:
        mock_cu.return_value = False

        with pytest.raises(errors.MissingOptionalComponents) as exc:
            report.to_feather(FEATHER_OUTPUT_PATH)
        assert (
            str(exc.value)
            == "some necessary libraries are not installed (hint: pip install pyarrow)"
        )


@pytest.mark.skipif(
    sys.version_info >= (3, 11, 0) or platform.python_implementation() != "CPython",
    reason="PyArrow does not support Python 3.11",
)
def test_to_parquet(report):
    report.to_parquet(str(PARQUET_OUTPUT_PATH))
    assert PARQUET_OUTPUT_PATH.is_file()

    df_csv = pd.read_csv(MOCK_CSV_PATH)
    df_csv["day"] = pd.to_datetime(df_csv["day"], format="%Y-%m-%d")
    df_parquet = pd.read_parquet(PARQUET_OUTPUT_PATH)
    assert df_parquet.equals(df_csv)

    os.remove(PARQUET_OUTPUT_PATH)


@pytest.mark.skipif(
    sys.version_info >= (3, 11, 0) or platform.python_implementation() != "CPython",
    reason="PyArrow does not support Python 3.11",
)
def test_to_parquet_no_extension(report):
    report.to_parquet(str(PARQUET_OUTPUT_PATH)[:-8])
    assert PARQUET_OUTPUT_PATH.is_file()

    df_csv = pd.read_csv(MOCK_CSV_PATH)
    df_csv["day"] = pd.to_datetime(df_csv["day"], format="%Y-%m-%d")
    df_parquet = pd.read_parquet(PARQUET_OUTPUT_PATH)
    assert df_parquet.equals(df_csv)

    os.remove(PARQUET_OUTPUT_PATH)


@pytest.mark.skipif(
    sys.version_info >= (3, 11, 0) or platform.python_implementation() != "CPython",
    reason="PyArrow does not support Python 3.11",
)
def test_to_parquet_no_pyarrow(report):
    with mock.patch.object(analytix, "can_use") as mock_cu:
        mock_cu.return_value = False

        with pytest.raises(errors.MissingOptionalComponents) as exc:
            report.to_parquet(PARQUET_OUTPUT_PATH)
        assert (
            str(exc.value)
            == "some necessary libraries are not installed (hint: pip install pyarrow)"
        )
