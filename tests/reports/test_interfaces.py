# SPDX-FileCopyrightText: 2021-2026 Ethan Henderson
# SPDX-FileCopyrightText: 2022 Jonxslays
# SPDX-License-Identifier: BSD-3-Clause

import datetime as dt
import json
import logging
from io import BytesIO
from io import StringIO
from unittest import mock

import pytest

from analytix import utils
from analytix.errors import DataFrameConversionError
from analytix.errors import MissingOptionalComponents
from analytix.reports.interfaces import Report
from analytix.reports.resources import ResultTable


def test_report_init(report: Report, report_type, result_table: ResultTable):
    assert report.resource == result_table
    assert report.type == report_type
    assert report.shape == (7, 2)


def test_reportshape_property(report: Report):
    assert report.shape == (7, 2)


def test_report_columns_property(report: Report):
    assert report.columns == ["day", "views"]


def test_report_dimensions_property(report: Report):
    assert report.dimensions == ["day"]


def test_report_metrics_property(report: Report):
    assert report.metrics == ["views"]


def test_report_to_json(report: Report, response_data: bytes):
    assert report.to_json() == response_data.decode("utf-8")


@pytest.mark.skipif(
    not logging.getLogger().isEnabledFor(logging.DEBUG),
    reason="DEBUG level logging is not enabled",
)
def test_report_to_json_to_path(report: Report, response_data: bytes, caplog):
    output = StringIO()

    with (
        mock.patch("analytix.reports.interfaces.Path.exists", return_value=False),
        mock.patch(
            "analytix.reports.interfaces.Path.write_text",
            lambda self, data: output.write(data),
        ),
        mock.patch(
            "analytix.reports.interfaces.Path.resolve",
            return_value="report.json",
        ),
    ):
        assert report.to_json("report.json") is None

    assert output.getvalue() == response_data.decode("utf-8")
    assert "Saved report to report.json" in caplog.text


def test_report_to_json_to_path_file_exists(report: Report):
    with (
        mock.patch("analytix.reports.interfaces.Path.exists", return_value=True),
        pytest.raises(FileExistsError) as exc_info,
    ):
        report.to_json("report.json")

    assert str(exc_info.value) == "file already exists and `overwrite` is set to False"


def test_report_to_json_to_buffer(report: Report, response_data: bytes):
    output = StringIO()
    assert report.to_json(output) is None
    assert output.getvalue() == response_data.decode("utf-8")


def test_report_to_json_to_invalid_object(report: Report):
    with pytest.raises(TypeError) as exc_info:
        report.to_json(123)

    assert str(exc_info.value) == "Expected str, PathLike, or TextIOBase, got int"


def test_report_to_csv(report: Report, report_csv: str):
    assert report.to_csv() == report_csv


@pytest.mark.skipif(
    not logging.getLogger().isEnabledFor(logging.DEBUG),
    reason="DEBUG level logging is not enabled",
)
def test_report_to_csv_to_path(report: Report, report_csv: str, caplog):
    output = StringIO()

    with (
        mock.patch("analytix.reports.interfaces.Path.exists", return_value=False),
        mock.patch(
            "analytix.reports.interfaces.Path.write_text",
            lambda self, data: output.write(data),
        ),
        mock.patch(
            "analytix.reports.interfaces.Path.resolve",
            return_value="report.csv",
        ),
    ):
        assert report.to_csv("report.csv") is None

    assert output.getvalue() == report_csv
    assert "Saved report to report.csv" in caplog.text


def test_report_to_csv_to_path_file_exists(report: Report):
    with (
        mock.patch("analytix.reports.interfaces.Path.exists", return_value=True),
        pytest.raises(FileExistsError) as exc_info,
    ):
        report.to_csv("report.csv")

    assert str(exc_info.value) == "file already exists and `overwrite` is set to False"


def test_report_to_csv_to_buffer(report: Report, report_csv: str):
    output = StringIO()
    assert report.to_csv(output) is None
    assert output.getvalue() == report_csv


def test_report_to_csv_to_invalid_object(report: Report):
    with pytest.raises(TypeError) as exc_info:
        report.to_csv(123)

    assert str(exc_info.value) == "Expected str, PathLike, or TextIOBase, got int"


def test_report_to_excel_no_openpyxl(report: Report):
    with (
        mock.patch("analytix.reports.interfaces.utils.can_use", return_value=False),
        pytest.raises(MissingOptionalComponents) as exc_info,
    ):
        report.to_excel("report.xlsx")

    assert (
        str(exc_info.value)
        == "some necessary libraries are not installed (hint: pip install openpyxl)"
    )


def test_report_to_excel_to_invalid_object(report: Report):
    with pytest.raises(TypeError) as exc_info:
        report.to_excel(123)

    assert str(exc_info.value) == "Expected str or PathLike, got int"


@pytest.mark.skipif(not utils.can_use("openpyxl"), reason="openpyxl is not installed")
@pytest.mark.skipif(
    not logging.getLogger().isEnabledFor(logging.DEBUG),
    reason="DEBUG level logging is not enabled",
)
def test_report_to_excel_create_new(report: Report):
    from openpyxl import Workbook

    output = BytesIO()

    class MockWorkbook(Workbook):
        instance: "MockWorkbook"

        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            MockWorkbook.instance = self

        def save(self, path):
            super().save(output)

    with (
        mock.patch("openpyxl.Workbook", MockWorkbook),
        mock.patch("analytix.reports.interfaces.Path.exists", return_value=False),
    ):
        report.to_excel("report.xlsx")

    assert output.getvalue().startswith(b"\x50\x4b\x03\x04")

    wb = MockWorkbook.instance
    assert len(wb.sheetnames) == 1
    assert wb.sheetnames[0] == "Analytics"


@pytest.mark.skipif(not utils.can_use("openpyxl"), reason="openpyxl is not installed")
@pytest.mark.skipif(
    not logging.getLogger().isEnabledFor(logging.DEBUG),
    reason="DEBUG level logging is not enabled",
)
def test_report_to_excel_append_to_existing(report: Report):
    from openpyxl import Workbook

    output = BytesIO()

    class MockWorkbook(Workbook):
        instance: "MockWorkbook"

        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            MockWorkbook.instance = self

        def save(self, path):
            super().save(output)

    with (
        mock.patch("openpyxl.load_workbook", return_value=MockWorkbook()),
        mock.patch("analytix.reports.interfaces.Path.exists", return_value=True),
    ):
        report.to_excel("report.xlsx")

    assert output.getvalue().startswith(b"\x50\x4b\x03\x04")

    wb = MockWorkbook.instance
    assert len(wb.sheetnames) == 2
    assert wb.sheetnames[1] == "Analytics"


def test_report_to_pandas_without_pandas(report: Report):
    with (
        mock.patch("analytix.reports.interfaces.utils.can_use", return_value=False),
        pytest.raises(MissingOptionalComponents) as exc_info,
    ):
        report.to_pandas()

    assert (
        str(exc_info.value)
        == "some necessary libraries are not installed (hint: pip install pandas)"
    )


@pytest.mark.skipif(not utils.can_use("pandas"), reason="pandas is not available")
def test_report_to_pandas_empty_df(empty_report: Report):
    assert empty_report.shape == (0, 2)

    with pytest.raises(DataFrameConversionError) as exc_info:
        empty_report.to_pandas()

    assert (
        str(exc_info.value)
        == "cannot convert to DataFrame as the returned data has no rows"
    )


@pytest.mark.skipif(not utils.can_use("pandas"), reason="pandas is not available")
def test_report_to_pandas(response_data, report: Report):
    import pandas as pd

    df = report.to_pandas()
    assert df.shape == (7, 2)
    assert list(df.columns) == report.columns

    assert df["day"][0] == pd.Timestamp(year=2022, month=6, day=20)
    for i, row in df.iterrows():
        assert list(row)[1:] == json.loads(response_data)["rows"][i][1:]


@mock.patch.object(utils, "can_use", return_value=False)
def test_report_to_polars_without_polars(_, report: Report):
    with pytest.raises(MissingOptionalComponents) as exc_info:
        report.to_polars()

    assert (
        str(exc_info.value)
        == "some necessary libraries are not installed (hint: pip install polars)"
    )


@pytest.mark.skipif(not utils.can_use("polars"), reason="polars is not available")
def test_report_to_polars_empty_df(empty_report: Report):
    assert empty_report.shape == (0, 2)

    with pytest.raises(DataFrameConversionError) as exc_info:
        empty_report.to_polars()

    assert (
        str(exc_info.value)
        == "cannot convert to DataFrame as the returned data has no rows"
    )


@pytest.mark.skipif(not utils.can_use("polars"), reason="Polars is not available")
def test_report_to_polars(response_data, report: Report):
    df = report.to_polars()
    assert df.shape == (7, 2)
    assert list(df.columns) == report.columns

    assert df["day"][0] == dt.date(2022, 6, 20)
    for i, row in enumerate(df.rows()):
        assert list(row)[1:] == json.loads(response_data)["rows"][i][1:]
